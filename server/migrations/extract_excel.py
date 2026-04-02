import openpyxl
import json

wb = openpyxl.load_workbook('Classes and Location Details as on 12-02-2026.xlsx')
ws = wb['Sheet1']

courses_by_dept = {}

for row_idx in range(5, ws.max_row + 1):
    class_name = ws.cell(row_idx, 2).value
    term_number = ws.cell(row_idx, 3).value
    course_full = ws.cell(row_idx, 4).value
    department = ws.cell(row_idx, 5).value
    school = ws.cell(row_idx, 6).value
    room_no = ws.cell(row_idx, 7).value
    block = ws.cell(row_idx, 8).value
    
    if not class_name or not department or not course_full:
        continue
    
    dept_clean = department.strip() if department else ''
    
    if dept_clean not in courses_by_dept:
        courses_by_dept[dept_clean] = {
            'school': school,
            'courses': {}
        }
    
    if course_full not in courses_by_dept[dept_clean]['courses']:
        courses_by_dept[dept_clean]['courses'][course_full] = []
    
    courses_by_dept[dept_clean]['courses'][course_full].append({
        'class': class_name,
        'term': term_number,
        'room': room_no,
        'block': block
    })

# Generate SQL INSERT statements with proper escaping
sql_inserts = []
for dept_name, dept_data in sorted(courses_by_dept.items()):
    courses_json = json.dumps(dept_data['courses'], ensure_ascii=False)
    school = dept_data['school'] or ''
    
    # Escape single quotes for PostgreSQL
    dept_name_escaped = dept_name.replace("'", "''")
    school_escaped = school.replace("'", "''")
    courses_escaped = courses_json.replace("'", "''")
    
    sql = f"INSERT INTO public.departments_courses (department_name, school, courses_json) VALUES (E'{dept_name_escaped}', E'{school_escaped}', '{courses_escaped}'::jsonb);"
    sql_inserts.append(sql)

# Write to file
with open('department_courses_insert.sql', 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_inserts))

print(f"Generated {len(sql_inserts)} INSERT statements")
print("File: department_courses_insert.sql")
